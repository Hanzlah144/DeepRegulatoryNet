import os
import logging
import pandas as pd
from analysis_pipeline import AnalysisPipeline
from mrna_overlap import overlap_mrnas

class SecondPipeline:
    def __init__(self, circ_file, mirna_file, deg_file, model_file, encoder_file, scaler_file, temp_dir, output_dir, data_dir=None):
        self.circ_file = circ_file
        self.mirna_file = mirna_file
        self.deg_file = deg_file
        self.model_file = model_file
        self.encoder_file = encoder_file
        self.scaler_file = scaler_file
        self.temp_dir = temp_dir
        self.output_dir = output_dir
        self.data_dir = data_dir

        os.makedirs(temp_dir, exist_ok=True)
        os.makedirs(output_dir, exist_ok=True)

        self.first_pipeline = AnalysisPipeline(
            circ_file=circ_file,
            mirna_file=mirna_file,
            deg_file=deg_file,
            model_file=model_file,
            encoder_file=encoder_file,
            scaler_file=scaler_file,
            temp_dir=temp_dir,
            output_dir=output_dir,
            data_dir=data_dir
        )

    def extract_overlapping_genes(self):
        logger = logging.getLogger()
        path = os.path.join(self.temp_dir, "overlapping_mrnas.csv")
        if not os.path.exists(path):
            logger.error("[ERROR]  %s not found", path)
            return []
        try:
            df = pd.read_csv(path)
            if "gene" not in df.columns:
                logger.error("[ERROR]  'gene' column missing in %s", path)
                return []
            genes = df["gene"].dropna().unique().tolist()
            output_path = os.path.join(self.output_dir, "overlapping_genes.csv")
            try:
                pd.DataFrame({"Gene": genes}).to_csv(output_path, index=False)
                logger.info("[INFO]  Saved: %s", output_path)
                logger.debug("[DEBUG] Extracted %d unique genes from %s", len(genes), path)
            except Exception as e:
                logger.error("[ERROR]  Failed to save %s: %s", output_path, e)
                return genes
            return genes
        except Exception as e:
            logger.error("[ERROR]  Failed to extract genes from %s: %s", path, e)
            return []

    def create_comprehensive_excel(self):
        
        
        
        
        logger = logging.getLogger()
        
        try:
            
            overlap_path = os.path.join(self.temp_dir, "overlapping_mrnas.csv")
            if not os.path.exists(overlap_path):
                logger.warning("[WARN]  No overlapping mRNAs file found, skipping Excel generation")
                return
            
            overlap_df = pd.read_csv(overlap_path)
            if overlap_df.empty:
                logger.warning("[WARN]  No overlapping data found, skipping Excel generation")
                return
            
            
            comprehensive_data = []
            
            
            import glob
            match_files = glob.glob(os.path.join(self.temp_dir, '*_strong_medium_matches.csv'))
            
            
            circ_mirna_map = {}
            logger.info(f"[INFO]  Found {len(match_files)} circRNA match files")
            
            for match_file in match_files:
                try:
                    match_df = pd.read_csv(match_file)
                    if 'circ_id' in match_df.columns and 'mirna_id' in match_df.columns:
                        circ_id = match_df['circ_id'].iloc[0] if not match_df.empty else os.path.basename(match_file).split('_')[0]
                        mirnas = match_df['mirna_id'].dropna().unique().tolist()
                        for mirna in mirnas:
                            if mirna not in circ_mirna_map:
                                circ_mirna_map[mirna] = []
                            circ_mirna_map[mirna].append(circ_id)
                        logger.debug(f"[DEBUG] {circ_id}: {len(mirnas)} miRNAs")
                except Exception as e:
                    logger.debug(f"[DEBUG] Error reading {match_file}: {e}")
                    continue
            
            logger.info(f"[INFO]  Created mapping for {len(circ_mirna_map)} miRNAs")
            
            
            logger.info(f"[INFO]  Processing {len(overlap_df)} gene-miRNA pairs")
            
            for _, row in overlap_df.iterrows():
                gene = row['gene']
                mirna = row['mirna']
                
                
                circrnas = circ_mirna_map.get(mirna, [])
                
                if circrnas:
                    
                    for circrna in circrnas:
                        comprehensive_data.append({
                            'Gene': gene,
                            'miRNA': mirna,
                            'circRNA': circrna,
                            'Interaction_Type': 'circRNA→miRNA→mRNA'
                        })
                else:
                    
                    comprehensive_data.append({
                        'Gene': gene,
                        'miRNA': mirna,
                        'circRNA': 'No circRNA interaction',
                        'Interaction_Type': 'miRNA→mRNA only'
                    })
            
            if not comprehensive_data:
                logger.warning("[WARN]  No comprehensive data generated")
                return
            
            
            comprehensive_df = pd.DataFrame(comprehensive_data)
            
            
            comprehensive_df = comprehensive_df.sort_values(['Gene', 'miRNA', 'circRNA'])
            
            
            excel_path = os.path.join(self.output_dir, "comprehensive_interactions.xlsx")
            
            
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                
                comprehensive_df.to_excel(writer, sheet_name='All_Interactions', index=False)
                
                
                summary_data = []
                for gene in comprehensive_df['Gene'].unique():
                    gene_data = comprehensive_df[comprehensive_df['Gene'] == gene]
                    mirnas = gene_data['miRNA'].unique()
                    circrnas = gene_data['circRNA'].unique()
                    circrnas = [c for c in circrnas if c != 'No circRNA interaction']
                    
                    summary_data.append({
                        'Gene': gene,
                        'Interacting_miRNAs': ', '.join(mirnas),
                        'Interacting_circRNAs': ', '.join(circrnas) if circrnas else 'None',
                        'Total_miRNAs': len(mirnas),
                        'Total_circRNAs': len(circrnas)
                    })
                
                summary_df = pd.DataFrame(summary_data)
                summary_df = summary_df.sort_values('Gene')
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                
                gene_centric = comprehensive_df.groupby('Gene').agg({
                    'miRNA': lambda x: ', '.join(x.unique()),
                    'circRNA': lambda x: ', '.join([c for c in x.unique() if c != 'No circRNA interaction'])
                }).reset_index()
                gene_centric.columns = ['Gene', 'All_Interacting_miRNAs', 'All_Interacting_circRNAs']
                gene_centric.to_excel(writer, sheet_name='Gene_Centric', index=False)
            
            
            try:
                from openpyxl import load_workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                
                
                wb = load_workbook(excel_path)
                
                
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    
                    
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    header_alignment = Alignment(horizontal="center", vertical="center")
                    
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                    
                    
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)  
                        ws.column_dimensions[column_letter].width = adjusted_width
                
                wb.save(excel_path)
                logger.info("[INFO]  Applied Excel formatting")
                
            except Exception as e:
                logger.debug(f"[DEBUG] Excel formatting failed: {e}")
                # Excel file is still saved, just without formatting
            
            logger.info("[INFO]  Saved comprehensive Excel: %s", excel_path)
            logger.info("[INFO]  Excel contains %d interaction records", len(comprehensive_df))
            
        except Exception as e:
            logger.error("[ERROR]  Failed to create comprehensive Excel: %s", e)

    def run(self):
        logger = logging.getLogger()
        for f in [self.circ_file, self.mirna_file, self.deg_file, self.model_file, self.encoder_file]:
            if not os.path.exists(f):
                raise FileNotFoundError(f"[ERROR] {f} not found")

        results = self.first_pipeline.process_all_circs()
        strong_hits, all_strong = self.first_pipeline.find_strong_hits(results)
        self.first_pipeline.match_mirnas(results, strong_hits, all_strong)
        overlapping = overlap_mrnas(self.deg_file, self.temp_dir, self.output_dir)
        self.extract_overlapping_genes()
        self.first_pipeline.construct_network(results, strong_hits, all_strong)
        
        
        self.create_comprehensive_excel()
        
        return results